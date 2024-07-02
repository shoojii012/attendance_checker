import csv
import os
import subprocess as sp
from datetime import timedelta

from celery import shared_task
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.db.models import Count
from django.template.loader import render_to_string
from django.utils import timezone

from .helper import (
    PingThreading,
    cumulative_time_overall,
    cumulative_time_this_month,
    current_users,
)
from .models import Device, Log, User


@shared_task
def check_attendance():
    devices = Device.objects.all()
    thread_list = []

    for i in range(2, 255):
        thread = PingThreading(ip_address=i)
        thread.start()
        thread_list.append(thread)

    for thread in thread_list:
        thread.join()

    output = sp.run(["arp", "-a"], capture_output=True, text=True).stdout
    now_time = timezone.now()
    active_users = {
        device.user for device in devices if device.mac_address.lower() in output.lower()
    }

    for user in User.objects.filter(is_active=True):
        if user in active_users:
            Log.objects.create(datetime=now_time, user=user)
            print(f"{user.name} entered at {now_time}")


@shared_task
def generate_statistics_html():
    context = {
        "monthly_ranking": cumulative_time_this_month(),
        "current_users": current_users(),
        "overall_ranking": cumulative_time_overall(),
    }
    html_content = render_to_string("statistics.html", context)
    output_path = os.path.join(settings.BASE_DIR, "attendance_checker", "templates", "index.html")

    with open(output_path, "w") as static_file:
        static_file.write(html_content)

    return output_path


@shared_task
def generate_monthly_report_csv():

    now = timezone.localtime()
    first_day_of_current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    first_day_of_previous_month = last_day_of_previous_month.replace(
        day=1, hour=0, minute=0, second=0, microsecond=0
    )

    # CSVファイルのパスを設定
    csv_filename = f"monthly_report_{last_day_of_previous_month.strftime('%Y%m')}.csv"
    csv_filepath = os.path.join(settings.MEDIA_ROOT, "reports", csv_filename)

    # ディレクトリが存在しない場合は作成
    os.makedirs(os.path.dirname(csv_filepath), exist_ok=True)

    # CSVファイルを書き込みモードでオープン
    with open(csv_filepath, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)

        # ヘッダーの書き込み
        headers = ["Name", "Date", "First Entry", "Last Exit", "Duration"]
        writer.writerow(headers)

        users = User.objects.filter(is_active=True)
        for user in users:
            logs = Log.objects.filter(
                user=user,
                datetime__gte=first_day_of_previous_month,
                datetime__lt=first_day_of_current_month,
            ).order_by("datetime")

            # 日ごとのログ数をカウント
            daily_log_counts = logs.values("datetime__date").annotate(count=Count("id"))
            daily_log_counts = {item["datetime__date"]: item["count"] for item in daily_log_counts}

            if logs.exists():
                date = None
                first_entry = None
                last_exit = None
                for log in logs:
                    log_date = timezone.localtime(log.datetime).date()
                    log_time = timezone.localtime(log.datetime).time()
                    if date != log_date:
                        if date:
                            log_count = daily_log_counts.get(date, 0)
                            duration = timedelta(minutes=log_count)
                            print(f"{user.name}:{date}:{duration.seconds}")
                            duration_str = f"{duration.seconds // 3600:02d}:{(duration.seconds % 3600) // 60:02d}:00"
                            writer.writerow(
                                [
                                    user.name,
                                    date.strftime("%Y/%m/%d"),
                                    first_entry.strftime("%H:%M"),
                                    last_exit.strftime("%H:%M"),
                                    duration_str,
                                ]
                            )
                        date = log_date
                        first_entry = log_time
                    last_exit = log_time

                # 最後の日のデータを書き込み
                if date:
                    log_count = daily_log_counts.get(date, 0)
                    duration = timedelta(minutes=log_count)
                    duration_str = (
                        f"{duration.seconds // 3600:02d}:{(duration.seconds % 3600) // 60:02d}:00"
                    )
                    writer.writerow(
                        [
                            user.name,
                            date.strftime("%Y/%m/%d"),
                            first_entry.strftime("%H:%M"),
                            last_exit.strftime("%H:%M"),
                            duration_str,
                        ]
                    )

    print(f"Monthly report generated and saved to {csv_filepath}")
    return csv_filepath


@shared_task
def generate_monthly_report_excel():
    now = timezone.now()
    first_day_of_current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    first_day_of_previous_month = last_day_of_previous_month.replace(
        day=1, hour=0, minute=0, second=0, microsecond=0
    )

    # Excelファイルのパスを設定
    excel_filename = f"monthly_report_{last_day_of_previous_month.strftime('%Y%m')}.xlsx"
    excel_filepath = os.path.join(settings.MEDIA_ROOT, "reports", excel_filename)

    # ディレクトリが存在しない場合は作成
    os.makedirs(os.path.dirname(excel_filepath), exist_ok=True)

    # ワークブックとシートの作成
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    # ヘッダーの設定
    headers = ["Name", "Date", "First Entry", "Last Exit", "Duration"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style="thin"))

    # データの書き込み
    row = 2
    users = User.objects.filter(is_active=True)
    for user in users:
        logs = Log.objects.filter(
            user=user,
            datetime__gte=first_day_of_previous_month,
            datetime__lt=first_day_of_current_month,
        ).order_by("datetime")

        # 日ごとのログ数をカウント
        daily_log_counts = logs.values("datetime__date").annotate(count=Count("id"))
        daily_log_counts = {item["datetime__date"]: item["count"] for item in daily_log_counts}

        if logs.exists():
            date = None
            first_entry = None
            last_exit = None
            for log in logs:
                log_date = log.datetime.date()
                log_time = log.datetime.time()
                if date != log_date:
                    if date:
                        log_count = daily_log_counts.get(date, 0)
                        duration = timedelta(minutes=log_count)
                        duration_str = f"{duration.seconds // 3600:02d}:{(duration.seconds % 3600) // 60:02d}:00"
                        ws.append(
                            [
                                user.username,
                                date.strftime("%Y/%m/%d"),
                                first_entry.strftime("%H:%M"),
                                last_exit.strftime("%H:%M"),
                                duration_str,
                            ]
                        )
                        row += 1
                    date = log_date
                    first_entry = log_time
                last_exit = log_time

            # 最後の日のデータを書き込み
            if date:
                log_count = daily_log_counts.get(date, 0)
                duration = timedelta(minutes=log_count)
                duration_str = (
                    f"{duration.seconds // 3600:02d}:{(duration.seconds % 3600) // 60:02d}:00"
                )
                ws.append(
                    [
                        user.username,
                        date.strftime("%Y/%m/%d"),
                        first_entry.strftime("%H:%M"),
                        last_exit.strftime("%H:%M"),
                        duration_str,
                    ]
                )
                row += 1

    # 列幅の自動調整
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # フィルターの追加
    ws.auto_filter.ref = f"A1:E{row-1}"

    # ファイルの保存
    wb.save(excel_filepath)
    print(f"Monthly report generated and saved to {excel_filepath}")
    return excel_filepath
